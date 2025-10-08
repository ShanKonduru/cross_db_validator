#!/usr/bin/env python3
"""
Create a new consolidated test suite Excel workbook with:
1. CONSOLIDATED_TESTS sheet - All test types in one sheet with TEST_TYPE column
2. REFERENCE sheet - Manual entry form with cascading dropdowns and color coding

This replaces the need for CONTROLLER sheet and multiple test type sheets.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import FormulaRule
from datetime import datetime

def create_consolidated_test_suite():
    """Create the new consolidated test suite workbook"""
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create CONSOLIDATED_TESTS sheet
    ws_tests = wb.create_sheet("CONSOLIDATED_TESTS")
    create_consolidated_tests_sheet(ws_tests)
    
    # Create REFERENCE sheet
    ws_ref = wb.create_sheet("REFERENCE")
    create_reference_sheet(ws_ref)
    
    # Save workbook
    filename = f"inputs/consolidated_test_suite_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    print(f"✅ Created new consolidated test suite: {filename}")
    
    return filename

def create_consolidated_tests_sheet(ws):
    """Create the consolidated tests sheet with all test types"""
    
    # Define column headers
    headers = [
        'Enable',              # MOVED TO FIRST - Enable/disable test execution
        'Test_Case_ID',
        'Test_Case_Name', 
        'TEST_TYPE',           # NEW: SMOKE, DATA_VALIDATION, CROSS_DB_VALIDATION
        'Test_Category',       # SETUP, CONNECTION, SCHEMA_VALIDATION, ROW_COUNT_VALIDATION, etc.
        'SRC_Application_Name',
        'SRC_Environment_Name',
        'SRC_Table_Name',      # NEW: Source table name
        'TGT_Application_Name',
        'TGT_Environment_Name',
        'TGT_Table_Name',      # NEW: Target table name
        'Priority',
        'Expected_Result',
        'Description',
        'Prerequisites',
        'Parameters',
        'Tags',
        'Expected_Duration_mins'
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    # Sample test cases covering all test types
    sample_tests = [
        # ================================
        # SMOKE TESTS - Connection & Basic Validation
        # ================================
        {
            'Enable': True,
            'Test_Case_ID': 'SMOKE_PG_001',
            'Test_Case_Name': 'PostgreSQL Connection Test - QA Environment',
            'TEST_TYPE': 'SMOKE',
            'Test_Category': 'CONNECTION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'QA',
            'SRC_Table_Name': '',
            'TGT_Application_Name': '',
            'TGT_Environment_Name': '',
            'TGT_Table_Name': '',
            'Priority': 'High',
            'Expected_Result': 'PASS',
            'Description': 'Test basic PostgreSQL database connectivity',
            'Prerequisites': 'Database server must be running',
            'Parameters': 'connection_timeout=30',
            'Tags': 'smoke,connection,postgresql',
            'Expected_Duration_mins': 1
        },
        {
            'Enable': True,
            'Test_Case_ID': 'SMOKE_SQL_001',
            'Test_Case_Name': 'SQL Server Connection Test - DEV Environment',
            'TEST_TYPE': 'SMOKE',
            'Test_Category': 'CONNECTION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'DEV',
            'SRC_Table_Name': '',
            'TGT_Application_Name': '',
            'TGT_Environment_Name': '',
            'TGT_Table_Name': '',
            'Priority': 'High',
            'Expected_Result': 'PASS',
            'Description': 'Test basic SQL Server database connectivity',
            'Prerequisites': 'SQL Server instance must be running',
            'Parameters': 'connection_timeout=45,retry_count=3',
            'Tags': 'smoke,connection,sqlserver',
            'Expected_Duration_mins': 2
        },
        {
            'Enable': True,
            'Test_Case_ID': 'SMOKE_OR_001',
            'Test_Case_Name': 'Oracle Connection Test - PROD Environment',
            'TEST_TYPE': 'SMOKE',
            'Test_Category': 'CONNECTION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'PROD',
            'SRC_Table_Name': '',
            'TGT_Application_Name': '',
            'TGT_Environment_Name': '',
            'TGT_Table_Name': '',
            'Priority': 'Critical',
            'Expected_Result': 'PASS',
            'Description': 'Test basic Oracle database connectivity',
            'Prerequisites': 'Oracle listener must be active',
            'Parameters': 'connection_timeout=60,pool_size=5',
            'Tags': 'smoke,connection,oracle',
            'Expected_Duration_mins': 1
        },
        {
            'Enable': True,
            'Test_Case_ID': 'SMOKE_TBL_001',
            'Test_Case_Name': 'Employee Table Existence Check',
            'TEST_TYPE': 'SMOKE',
            'Test_Category': 'TABLE_EXISTS',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'QA',
            'SRC_Table_Name': 'public.employees',
            'TGT_Application_Name': '',
            'TGT_Environment_Name': '',
            'TGT_Table_Name': '',
            'Priority': 'High',
            'Expected_Result': 'PASS',
            'Description': 'Verify that employees table exists and is accessible',
            'Prerequisites': 'Database connection established',
            'Parameters': '',
            'Tags': 'smoke,table,employees',
            'Expected_Duration_mins': 1
        },
        {
            'Enable': True,
            'Test_Case_ID': 'SMOKE_TBL_002',
            'Test_Case_Name': 'Orders Table Existence Check',
            'TEST_TYPE': 'SMOKE',
            'Test_Category': 'TABLE_EXISTS',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'DEV',
            'SRC_Table_Name': 'dbo.orders',
            'TGT_Application_Name': '',
            'TGT_Environment_Name': '',
            'TGT_Table_Name': '',
            'Priority': 'Medium',
            'Expected_Result': 'PASS',
            'Description': 'Verify that orders table exists and is accessible',
            'Prerequisites': 'Database connection established',
            'Parameters': '',
            'Tags': 'smoke,table,orders',
            'Expected_Duration_mins': 1
        },
        {
            'Enable': True,
            'Test_Case_ID': 'SMOKE_SEL_001',
            'Test_Case_Name': 'Basic Table Select Test',
            'TEST_TYPE': 'SMOKE',
            'Test_Category': 'TABLE_SELECT',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'QA',
            'SRC_Table_Name': 'public.employees',
            'TGT_Application_Name': '',
            'TGT_Environment_Name': '',
            'TGT_Table_Name': '',
            'Priority': 'Medium',
            'Expected_Result': 'PASS',
            'Description': 'Test basic SELECT operation on employees table',
            'Prerequisites': 'Table exists and has data',
            'Parameters': 'limit=10',
            'Tags': 'smoke,select,employees',
            'Expected_Duration_mins': 1
        },
        
        # ================================
        # DATA_VALIDATION TESTS - Various Tolerance Scenarios
        # ================================
        {
            'Enable': True,
            'Test_Case_ID': 'DVAL_RC_001',
            'Test_Case_Name': 'Employee Row Count - Exact Match',
            'TEST_TYPE': 'DATA_VALIDATION',
            'Test_Category': 'ROW_COUNT_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'QA',
            'SRC_Table_Name': 'public.employees',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'DEV',
            'TGT_Table_Name': 'public.employees',
            'Priority': 'High',
            'Expected_Result': 'PASS',
            'Description': 'Validate exact row count match between environments',
            'Prerequisites': 'Both databases accessible',
            'Parameters': '',
            'Tags': 'data_validation,row_count,exact_match',
            'Expected_Duration_mins': 2
        },
        {
            'Enable': True,
            'Test_Case_ID': 'DVAL_RC_002',
            'Test_Case_Name': 'Employee Row Count - 5% Tolerance',
            'TEST_TYPE': 'DATA_VALIDATION',
            'Test_Category': 'ROW_COUNT_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'QA',
            'SRC_Table_Name': 'public.employees',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'UAT',
            'TGT_Table_Name': 'public.employees',
            'Priority': 'Medium',
            'Expected_Result': 'PASS',
            'Description': 'Validate row count with 5% percentage tolerance',
            'Prerequisites': 'Both databases accessible',
            'Parameters': 'tolerance=5%,tolerance_type=percentage',
            'Tags': 'data_validation,row_count,percentage_tolerance',
            'Expected_Duration_mins': 2
        },
        {
            'Enable': True,
            'Test_Case_ID': 'DVAL_RC_003',
            'Test_Case_Name': 'Orders Row Count - 10% Tolerance',
            'TEST_TYPE': 'DATA_VALIDATION',
            'Test_Category': 'ROW_COUNT_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'DEV',
            'SRC_Table_Name': 'dbo.orders',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'UAT',
            'TGT_Table_Name': 'dbo.orders',
            'Priority': 'Medium',
            'Expected_Result': 'PASS',
            'Description': 'Validate orders row count with 10% percentage tolerance',
            'Prerequisites': 'Both databases accessible',
            'Parameters': 'tolerance=10%,tolerance_type=percentage',
            'Tags': 'data_validation,row_count,orders',
            'Expected_Duration_mins': 3
        },
        {
            'Enable': True,
            'Test_Case_ID': 'DVAL_RC_004',
            'Test_Case_Name': 'Products Row Count - 100 Absolute Tolerance',
            'TEST_TYPE': 'DATA_VALIDATION',
            'Test_Category': 'ROW_COUNT_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'QA',
            'SRC_Table_Name': 'public.products',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'PROD',
            'TGT_Table_Name': 'public.products',
            'Priority': 'Low',
            'Expected_Result': 'PASS',
            'Description': 'Validate products row count with absolute tolerance of 100 rows',
            'Prerequisites': 'Both databases accessible',
            'Parameters': 'tolerance=100,tolerance_type=absolute',
            'Tags': 'data_validation,row_count,absolute_tolerance',
            'Expected_Duration_mins': 2
        },
        {
            'Enable': True,
            'Test_Case_ID': 'DVAL_RC_005',
            'Test_Case_Name': 'Customer Row Count - 50 Absolute Tolerance',
            'TEST_TYPE': 'DATA_VALIDATION',
            'Test_Category': 'ROW_COUNT_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'DEV',
            'SRC_Table_Name': 'dbo.customers',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'ACC',
            'TGT_Table_Name': 'dbo.customers',
            'Priority': 'Medium',
            'Expected_Result': 'PASS',
            'Description': 'Validate customer row count with absolute tolerance of 50 rows',
            'Prerequisites': 'Both databases accessible',
            'Parameters': 'tolerance=50,tolerance_type=absolute',
            'Tags': 'data_validation,row_count,customers',
            'Expected_Duration_mins': 2
        },
        {
            'Enable': True,
            'Test_Case_ID': 'DVAL_SC_001',
            'Test_Case_Name': 'Employee Schema - Exact Match',
            'TEST_TYPE': 'DATA_VALIDATION',
            'Test_Category': 'SCHEMA_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'QA',
            'SRC_Table_Name': 'public.employees',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'DEV',
            'TGT_Table_Name': 'public.employees',
            'Priority': 'High',
            'Expected_Result': 'PASS',
            'Description': 'Validate exact schema structure match between environments',
            'Prerequisites': 'Both databases accessible',
            'Parameters': '',
            'Tags': 'data_validation,schema,exact_match',
            'Expected_Duration_mins': 3
        },
        {
            'Enable': True,
            'Test_Case_ID': 'DVAL_SC_002',
            'Test_Case_Name': 'Orders Schema - Column Subset Validation',
            'TEST_TYPE': 'DATA_VALIDATION',
            'Test_Category': 'SCHEMA_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'DEV',
            'SRC_Table_Name': 'dbo.orders',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'UAT',
            'TGT_Table_Name': 'dbo.orders',
            'Priority': 'Medium',
            'Expected_Result': 'PASS',
            'Description': 'Validate specific columns exist in target schema',
            'Prerequisites': 'Both databases accessible',
            'Parameters': 'validate_columns=order_id|customer_id|order_date|total_amount',
            'Tags': 'data_validation,schema,subset',
            'Expected_Duration_mins': 3
        },
        {
            'Enable': True,
            'Test_Case_ID': 'DVAL_SC_003',
            'Test_Case_Name': 'Products Schema - Data Type Validation',
            'TEST_TYPE': 'DATA_VALIDATION',
            'Test_Category': 'SCHEMA_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'QA',
            'SRC_Table_Name': 'public.products',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'PROD',
            'TGT_Table_Name': 'public.products',
            'Priority': 'High',
            'Expected_Result': 'PASS',
            'Description': 'Validate data types match between environments',
            'Prerequisites': 'Both databases accessible',
            'Parameters': 'validate_datatypes=true,ignore_nullable=false',
            'Tags': 'data_validation,schema,datatypes',
            'Expected_Duration_mins': 4
        },
        
        # ================================
        # CROSS_DB_VALIDATION TESTS - Column Comparisons & Complex Scenarios
        # ================================
        {
            'Enable': True,
            'Test_Case_ID': 'CROSS_CC_001',
            'Test_Case_Name': 'Employee Data - Exact Column Match',
            'TEST_TYPE': 'CROSS_DB_VALIDATION',
            'Test_Category': 'COL_COL_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'QA',
            'SRC_Table_Name': 'public.employees',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'DEV',
            'TGT_Table_Name': 'public.employees',
            'Priority': 'High',
            'Expected_Result': 'PASS',
            'Description': 'Compare employee data with exact column matching',
            'Prerequisites': 'Both databases accessible and contain employee data',
            'Parameters': 'compare_columns=emp_id|first_name|last_name|email,key_column=emp_id',
            'Tags': 'cross_db,column_validation,exact_match',
            'Expected_Duration_mins': 5
        },
        {
            'Enable': True,
            'Test_Case_ID': 'CROSS_CC_002',
            'Test_Case_Name': 'Employee Data - Case Insensitive Match',
            'TEST_TYPE': 'CROSS_DB_VALIDATION',
            'Test_Category': 'COL_COL_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'DEV',
            'SRC_Table_Name': 'public.employees',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'UAT',
            'TGT_Table_Name': 'public.employees',
            'Priority': 'Medium',
            'Expected_Result': 'PASS',
            'Description': 'Compare employee data with case insensitive matching',
            'Prerequisites': 'Both databases accessible',
            'Parameters': 'compare_columns=first_name|last_name,key_column=emp_id,case_sensitive=false',
            'Tags': 'cross_db,column_validation,case_insensitive',
            'Expected_Duration_mins': 4
        },
        {
            'Enable': True,
            'Test_Case_ID': 'CROSS_CC_003',
            'Test_Case_Name': 'Orders Data - Numeric Tolerance',
            'TEST_TYPE': 'CROSS_DB_VALIDATION',
            'Test_Category': 'COL_COL_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'QA',
            'SRC_Table_Name': 'dbo.orders',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'DEV',
            'TGT_Table_Name': 'dbo.orders',
            'Priority': 'Medium',
            'Expected_Result': 'PASS',
            'Description': 'Compare order amounts with numeric tolerance',
            'Prerequisites': 'Both databases accessible',
            'Parameters': 'compare_columns=total_amount,key_column=order_id,numeric_tolerance=0.01',
            'Tags': 'cross_db,column_validation,numeric_tolerance',
            'Expected_Duration_mins': 6
        },
        {
            'Enable': True,
            'Test_Case_ID': 'CROSS_CC_004',
            'Test_Case_Name': 'Products Data - Partial Column Set',
            'TEST_TYPE': 'CROSS_DB_VALIDATION',
            'Test_Category': 'COL_COL_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'DEV',
            'SRC_Table_Name': 'public.products',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'PROD',
            'TGT_Table_Name': 'public.products',
            'Priority': 'Low',
            'Expected_Result': 'PASS',
            'Description': 'Compare subset of product columns',
            'Prerequisites': 'Both databases accessible',
            'Parameters': 'compare_columns=product_name|price,key_column=product_id,allow_nulls=true',
            'Tags': 'cross_db,column_validation,subset',
            'Expected_Duration_mins': 4
        },
        {
            'Enable': True,
            'Test_Case_ID': 'CROSS_CC_005',
            'Test_Case_Name': 'Financial Data - Expected Differences with Tolerance',
            'TEST_TYPE': 'CROSS_DB_VALIDATION',
            'Test_Category': 'COL_COL_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'QA',
            'SRC_Table_Name': 'dbo.transactions',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'DEV',
            'TGT_Table_Name': 'dbo.transactions',
            'Priority': 'High',
            'Expected_Result': 'PASS',
            'Description': 'Compare financial columns with expected differences in tax calculations',
            'Prerequisites': 'Both databases accessible, tax calculations may differ',
            'Parameters': 'compare_columns=transaction_id|amount|currency,expect_cols=tax_amount|total_amount,key_column=transaction_id,numeric_tolerance=0.05',
            'Tags': 'cross_db,column_validation,expected_differences,financial',
            'Expected_Duration_mins': 7
        },
        {
            'Enable': True,
            'Test_Case_ID': 'CROSS_CC_006',
            'Test_Case_Name': 'Employee Salary - Expected Tolerance on Bonus',
            'TEST_TYPE': 'CROSS_DB_VALIDATION',
            'Test_Category': 'COL_COL_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'QA',
            'SRC_Table_Name': 'public.employees',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'PROD',
            'TGT_Table_Name': 'public.employees',
            'Priority': 'Medium',
            'Expected_Result': 'PASS',
            'Description': 'Compare employee data expecting differences in bonus calculations',
            'Prerequisites': 'Bonus calculations may vary between environments',
            'Parameters': 'compare_columns=emp_id|first_name|last_name|base_salary,expect_cols=bonus|total_salary,key_column=emp_id,numeric_tolerance=100.00',
            'Tags': 'cross_db,column_validation,salary,bonus_tolerance',
            'Expected_Duration_mins': 6
        },
        {
            'Enable': True,
            'Test_Case_ID': 'CROSS_CC_007',
            'Test_Case_Name': 'Inventory - Stock Level Tolerance',
            'TEST_TYPE': 'CROSS_DB_VALIDATION',
            'Test_Category': 'COL_COL_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'DEV',
            'SRC_Table_Name': 'dbo.inventory',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'UAT',
            'TGT_Table_Name': 'dbo.inventory',
            'Priority': 'Medium',
            'Expected_Result': 'PASS',
            'Description': 'Compare inventory with expected stock level differences',
            'Prerequisites': 'Stock levels may differ due to real-time updates',
            'Parameters': 'compare_columns=product_id|product_name|location,expect_cols=stock_quantity|reserved_quantity,key_column=product_id,numeric_tolerance=10,tolerance_type=absolute',
            'Tags': 'cross_db,column_validation,inventory,stock_tolerance',
            'Expected_Duration_mins': 5
        },
        {
            'Enable': True,
            'Test_Case_ID': 'CROSS_CC_008',
            'Test_Case_Name': 'Orders - Timestamp and Status Exclusions',
            'TEST_TYPE': 'CROSS_DB_VALIDATION',
            'Test_Category': 'COL_COL_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'QA',
            'SRC_Table_Name': 'dbo.orders',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'ACC',
            'TGT_Table_Name': 'dbo.orders',
            'Priority': 'High',
            'Expected_Result': 'PASS',
            'Description': 'Compare order data excluding timestamp and status columns that are expected to differ',
            'Prerequisites': 'Timestamps and statuses may vary between environments',
            'Parameters': 'compare_columns=order_id|customer_id|order_date|total_amount,expect_cols=created_timestamp|updated_timestamp|status|processing_notes,key_column=order_id',
            'Tags': 'cross_db,column_validation,timestamp_exclusion',
            'Expected_Duration_mins': 6
        },
        {
            'Enable': True,
            'Test_Case_ID': 'CROSS_CC_009',
            'Test_Case_Name': 'Customer Data - Percentage Tolerance on Metrics',
            'TEST_TYPE': 'CROSS_DB_VALIDATION',
            'Test_Category': 'COL_COL_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'DEV',
            'SRC_Table_Name': 'dbo.customers',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'PROD',
            'TGT_Table_Name': 'dbo.customers',
            'Priority': 'Medium',
            'Expected_Result': 'PASS',
            'Description': 'Compare customer data with percentage tolerance on calculated metrics',
            'Prerequisites': 'Customer metrics calculations may have minor variations',
            'Parameters': 'compare_columns=customer_id|first_name|last_name|email,expect_cols=lifetime_value|credit_score|loyalty_points,key_column=customer_id,numeric_tolerance=5%,tolerance_type=percentage',
            'Tags': 'cross_db,column_validation,customer_metrics,percentage_tolerance',
            'Expected_Duration_mins': 8
        },
        {
            'Enable': True,
            'Test_Case_ID': 'CROSS_CC_010',
            'Test_Case_Name': 'Product Pricing - Multi-Currency Tolerance',
            'TEST_TYPE': 'CROSS_DB_VALIDATION',
            'Test_Category': 'COL_COL_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'QA',
            'SRC_Table_Name': 'public.products',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'UAT',
            'TGT_Table_Name': 'public.products',
            'Priority': 'High',
            'Expected_Result': 'PASS',
            'Description': 'Compare product data with expected differences in currency-converted prices',
            'Prerequisites': 'Currency exchange rates may cause price variations',
            'Parameters': 'compare_columns=product_id|product_name|category|base_price_usd,expect_cols=price_eur|price_gbp|price_jpy,key_column=product_id,numeric_tolerance=2%,tolerance_type=percentage,currency_conversion=true',
            'Tags': 'cross_db,column_validation,multi_currency,exchange_rate_tolerance',
            'Expected_Duration_mins': 9
        },
        {
            'Enable': True,
            'Test_Case_ID': 'CROSS_CC_011',
            'Test_Case_Name': 'Analytics Data - Calculated Fields with Tolerance',
            'TEST_TYPE': 'CROSS_DB_VALIDATION',
            'Test_Category': 'COL_COL_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'DEV',
            'SRC_Table_Name': 'dbo.analytics_summary',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'ACC',
            'TGT_Table_Name': 'dbo.analytics_summary',
            'Priority': 'Medium',
            'Expected_Result': 'PASS',
            'Description': 'Compare analytics data with tolerance on calculated aggregations',
            'Prerequisites': 'Aggregation calculations may have rounding differences',
            'Parameters': 'compare_columns=summary_id|date|category,expect_cols=total_revenue|avg_order_value|conversion_rate|growth_percentage,key_column=summary_id,numeric_tolerance=0.001,decimal_precision=4',
            'Tags': 'cross_db,column_validation,analytics,aggregation_tolerance',
            'Expected_Duration_mins': 10
        },
        {
            'Enable': True,
            'Test_Case_ID': 'CROSS_RC_001',
            'Test_Case_Name': 'Employee Row Count - Cross DB Exact',
            'TEST_TYPE': 'CROSS_DB_VALIDATION',
            'Test_Category': 'ROW_COUNT_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'QA',
            'SRC_Table_Name': 'public.employees',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'DEV',
            'TGT_Table_Name': 'public.employees',
            'Priority': 'High',
            'Expected_Result': 'PASS',
            'Description': 'Cross-database row count validation without tolerance',
            'Prerequisites': 'Both databases accessible',
            'Parameters': '',
            'Tags': 'cross_db,row_count,exact',
            'Expected_Duration_mins': 3
        },
        {
            'Enable': True,
            'Test_Case_ID': 'CROSS_RC_002',
            'Test_Case_Name': 'Orders Row Count - Cross DB 2% Tolerance',
            'TEST_TYPE': 'CROSS_DB_VALIDATION',
            'Test_Category': 'ROW_COUNT_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'DEV',
            'SRC_Table_Name': 'dbo.orders',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'UAT',
            'TGT_Table_Name': 'dbo.orders',
            'Priority': 'Medium',
            'Expected_Result': 'PASS',
            'Description': 'Cross-database row count with 2% tolerance',
            'Prerequisites': 'Both databases accessible',
            'Parameters': 'tolerance=2%,tolerance_type=percentage',
            'Tags': 'cross_db,row_count,percentage_tolerance',
            'Expected_Duration_mins': 3
        },
        {
            'Enable': True,
            'Test_Case_ID': 'CROSS_RC_003',
            'Test_Case_Name': 'Customer Row Count - Cross DB 25 Absolute',
            'TEST_TYPE': 'CROSS_DB_VALIDATION',
            'Test_Category': 'ROW_COUNT_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'QA',
            'SRC_Table_Name': 'dbo.customers',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'PROD',
            'TGT_Table_Name': 'dbo.customers',
            'Priority': 'Low',
            'Expected_Result': 'PASS',
            'Description': 'Cross-database row count with absolute tolerance of 25',
            'Prerequisites': 'Both databases accessible',
            'Parameters': 'tolerance=25,tolerance_type=absolute',
            'Tags': 'cross_db,row_count,absolute_tolerance',
            'Expected_Duration_mins': 3
        },
        
        # ================================
        # COMPLEX SCENARIOS - Mixed Parameters & Edge Cases
        # ================================
        {
            'Enable': True,
            'Test_Case_ID': 'COMPLEX_001',
            'Test_Case_Name': 'Multi-Environment Employee Sync',
            'TEST_TYPE': 'CROSS_DB_VALIDATION',
            'Test_Category': 'COL_COL_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'QA',
            'SRC_Table_Name': 'public.employees',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'PROD',
            'TGT_Table_Name': 'public.employees',
            'Priority': 'Critical',
            'Expected_Result': 'PASS',
            'Description': 'Complex validation with multiple parameters',
            'Prerequisites': 'QA and PROD environments must be in sync',
            'Parameters': 'compare_columns=emp_id|email|department,key_column=emp_id,case_sensitive=false,allow_nulls=false,trim_spaces=true',
            'Tags': 'cross_db,complex,multi_param',
            'Expected_Duration_mins': 8
        },
        {
            'Enable': True,
            'Test_Case_ID': 'COMPLEX_002',
            'Test_Case_Name': 'Financial Data Precision Validation',
            'TEST_TYPE': 'CROSS_DB_VALIDATION',
            'Test_Category': 'COL_COL_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'DEV',
            'SRC_Table_Name': 'dbo.transactions',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'ACC',
            'TGT_Table_Name': 'dbo.transactions',
            'Priority': 'Critical',
            'Expected_Result': 'PASS',
            'Description': 'High precision financial data comparison',
            'Prerequisites': 'Financial data must be precise',
            'Parameters': 'compare_columns=amount|tax|total,key_column=transaction_id,numeric_tolerance=0.001,decimal_precision=3',
            'Tags': 'cross_db,financial,precision',
            'Expected_Duration_mins': 10
        },
        {
            'Enable': False,
            'Test_Case_ID': 'DISABLED_001',
            'Test_Case_Name': 'Disabled Test Example',
            'TEST_TYPE': 'SMOKE',
            'Test_Category': 'CONNECTION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'TEST',
            'SRC_Table_Name': '',
            'TGT_Application_Name': '',
            'TGT_Environment_Name': '',
            'TGT_Table_Name': '',
            'Priority': 'Low',
            'Expected_Result': 'PASS',
            'Description': 'Example of a disabled test case',
            'Prerequisites': 'Test environment setup',
            'Parameters': 'connection_timeout=30',
            'Tags': 'disabled,example',
            'Expected_Duration_mins': 1
        },
        {
            'Enable': True,
            'Test_Case_ID': 'EDGE_001',
            'Test_Case_Name': 'Large Table Row Count - 1% Tolerance',
            'TEST_TYPE': 'DATA_VALIDATION',
            'Test_Category': 'ROW_COUNT_VALIDATION',
            'SRC_Application_Name': 'DUMMY',
            'SRC_Environment_Name': 'QA',
            'SRC_Table_Name': 'public.big_table',
            'TGT_Application_Name': 'DUMMY',
            'TGT_Environment_Name': 'DEV',
            'TGT_Table_Name': 'public.big_table',
            'Priority': 'Medium',
            'Expected_Result': 'PASS',
            'Description': 'Row count validation for large tables with minimal tolerance',
            'Prerequisites': 'Large datasets available',
            'Parameters': 'tolerance=1%,tolerance_type=percentage,batch_size=10000',
            'Tags': 'data_validation,large_table,performance',
            'Expected_Duration_mins': 15
        }
    ]
    
    # Add sample test data
    for row, test in enumerate(sample_tests, 2):
        # Determine the color for this test type
        test_type = test.get('TEST_TYPE', '')
        row_color = None
        if test_type == 'SMOKE':
            row_color = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
        elif test_type == 'DATA_VALIDATION':
            row_color = PatternFill(start_color="E8F0FF", end_color="E8F0FF", fill_type="solid")
        elif test_type == 'CROSS_DB_VALIDATION':
            row_color = PatternFill(start_color="FFF0E8", end_color="FFF0E8", fill_type="solid")
        
        # Apply data and formatting to all columns in the row
        for col, header in enumerate(headers, 1):
            value = test.get(header, '')
            cell = ws.cell(row=row, column=col, value=value)
            
            # Apply the row color to entire row
            if row_color:
                cell.fill = row_color
            
            # Apply border
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
    # Auto-fit columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    
    print("✅ Created CONSOLIDATED_TESTS sheet with sample data")

def create_reference_sheet(ws):
    """Create the reference sheet for manual test case entry"""
    
    # Title
    ws.merge_cells('A1:V1')
    title_cell = ws['A1']
    title_cell.value = "📋 Manual Test Case Entry Form"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Instructions
    ws.merge_cells('A3:V3')
    instruction_cell = ws['A3']
    instruction_cell.value = "Use this form to manually add new test cases. Dropdowns provide valid options with color coding."
    instruction_cell.font = Font(italic=True, color="666666")
    instruction_cell.alignment = Alignment(horizontal="center")
    
    # Form fields starting from row 5
    form_fields = [
        ('Enable', 'A5', 'Enable test (dropdown)'),
        ('Test_Case_ID', 'A6', 'Enter unique test ID (e.g., SMOKE_PG_001)'),
        ('Test_Case_Name', 'A7', 'Descriptive name for the test case'),
        ('TEST_TYPE', 'A8', 'Select test type (dropdown)'),
        ('Test_Category', 'A9', 'Select test category (dropdown)'),
        ('SRC_Application_Name', 'A10', 'Source application name'),
        ('SRC_Environment_Name', 'A11', 'Source environment (dropdown)'),
        ('SRC_Table_Name', 'A12', 'Source table name'),
        ('TGT_Application_Name', 'A13', 'Target application name'),
        ('TGT_Environment_Name', 'A14', 'Target environment (dropdown)'),
        ('TGT_Table_Name', 'A15', 'Target table name'),
        ('Priority', 'A16', 'Test priority (dropdown)'),
        ('Expected_Result', 'A17', 'Expected result (dropdown)'),
        ('Description', 'A18', 'Detailed test description'),
        ('Prerequisites', 'A19', 'Test prerequisites'),
        ('Parameters', 'A20', 'Test parameters (key=value pairs)'),
        ('Tags', 'A21', 'Comma-separated tags'),
        ('Expected_Duration_mins', 'A22', 'Expected duration in minutes')
    ]
    
    # Create form
    for field_name, cell_addr, placeholder in form_fields:
        # Label
        label_cell = ws[cell_addr]
        label_cell.value = field_name
        label_cell.font = Font(bold=True)
        label_cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        # Input cell (next column)
        input_col = chr(ord(cell_addr[0]) + 2)  # Skip one column for spacing
        input_cell_addr = f"{input_col}{cell_addr[1:]}"
        input_cell = ws[input_cell_addr]
        input_cell.value = placeholder
        input_cell.font = Font(italic=True, color="999999")
        
        # Add borders
        for cell in [label_cell, input_cell]:
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
    # Create dropdown validations
    create_dropdown_validations(ws)
    
    # Reference data section
    create_reference_data_section(ws)
    
    print("✅ Created REFERENCE sheet with manual entry form")

def create_dropdown_validations(ws):
    """Create dropdown validations for the reference sheet"""
    
    # Enable dropdown (now first field)
    enable_validation = DataValidation(
        type="list",
        formula1='"TRUE,FALSE"',
        allow_blank=False
    )
    enable_validation.error = "Please select TRUE or FALSE"
    enable_validation.errorTitle = "Invalid Enable Value"
    ws.add_data_validation(enable_validation)
    enable_validation.add('C5')  # Enable field
    
    # TEST_TYPE dropdown
    test_type_validation = DataValidation(
        type="list",
        formula1='"SMOKE,DATA_VALIDATION,CROSS_DB_VALIDATION"',
        allow_blank=False
    )
    test_type_validation.error = "Please select a valid test type"
    test_type_validation.errorTitle = "Invalid Test Type"
    ws.add_data_validation(test_type_validation)
    test_type_validation.add('C8')
    
    # Test_Category dropdown (cascading based on TEST_TYPE)
    category_validation = DataValidation(
        type="list",
        formula1='"CONNECTION,TABLE_EXISTS,TABLE_SELECT,SETUP,SCHEMA_VALIDATION,ROW_COUNT_VALIDATION,COL_COL_VALIDATION"',
        allow_blank=False
    )
    category_validation.error = "Please select a valid test category"
    category_validation.errorTitle = "Invalid Test Category"
    ws.add_data_validation(category_validation)
    category_validation.add('C9')
    
    # Environment dropdown
    env_validation = DataValidation(
        type="list",
        formula1='"QA,DEV,UAT,NP1,ACC,PROD"',
        allow_blank=False
    )
    env_validation.error = "Please select a valid environment"
    env_validation.errorTitle = "Invalid Environment"
    ws.add_data_validation(env_validation)
    env_validation.add('C11')  # SRC_Environment
    env_validation.add('C14')  # TGT_Environment
    
    # Priority dropdown
    priority_validation = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        allow_blank=False
    )
    priority_validation.error = "Please select a valid priority"
    priority_validation.errorTitle = "Invalid Priority"
    ws.add_data_validation(priority_validation)
    priority_validation.add('C16')
    
    # Expected_Result dropdown
    result_validation = DataValidation(
        type="list",
        formula1='"PASS,FAIL"',
        allow_blank=False
    )
    result_validation.error = "Please select PASS or FAIL"
    result_validation.errorTitle = "Invalid Expected Result"
    ws.add_data_validation(result_validation)
    result_validation.add('C17')

def create_reference_data_section(ws):
    """Create reference data section with valid values and color coding"""
    
    # Reference data starting from row 25
    ws.merge_cells('A25:V25')
    ref_title = ws['A25']
    ref_title.value = "📚 Reference Data & Color Coding"
    ref_title.font = Font(bold=True, size=14, color="FFFFFF")
    ref_title.fill = PatternFill(start_color="5D4E75", end_color="5D4E75", fill_type="solid")
    ref_title.alignment = Alignment(horizontal="center")
    
    # TEST_TYPE reference
    ws['A27'].value = "TEST_TYPE Values:"
    ws['A27'].font = Font(bold=True)
    
    test_types = [
        ('SMOKE', 'Connection and basic functionality tests', 'E8F5E8'),
        ('DATA_VALIDATION', 'Data consistency and validation tests', 'E8F0FF'),
        ('CROSS_DB_VALIDATION', 'Cross-database comparison tests', 'FFF0E8')
    ]
    
    for i, (test_type, description, color) in enumerate(test_types, 28):
        ws[f'A{i}'].value = test_type
        ws[f'B{i}'].value = description
        ws[f'A{i}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f'B{i}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    # Test_Category reference
    ws['D27'].value = "Test_Category Values:"
    ws['D27'].font = Font(bold=True)
    
    categories = [
        ('CONNECTION', 'Database connectivity tests'),
        ('TABLE_EXISTS', 'Table existence verification'),
        ('TABLE_SELECT', 'Table accessibility tests'),
        ('SETUP', 'Environment setup tests'),
        ('SCHEMA_VALIDATION', 'Schema structure validation'),
        ('ROW_COUNT_VALIDATION', 'Row count comparison'),
        ('COL_COL_VALIDATION', 'Column-to-column comparison')
    ]
    
    for i, (category, description) in enumerate(categories, 28):
        ws[f'D{i}'].value = category
        ws[f'E{i}'].value = description
    
    # Environment reference
    ws['G27'].value = "Environment Values:"
    ws['G27'].font = Font(bold=True)
    
    environments = ['QA', 'DEV', 'UAT', 'NP1', 'ACC', 'PROD']
    for i, env in enumerate(environments, 28):
        ws[f'G{i}'].value = env
    
    # Auto-fit columns (skip merged cells)
    for col_num in range(1, 8):  # Adjust to actual number of columns used
        column_letter = chr(ord('A') + col_num - 1)
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell = ws[f'{column_letter}{row}']
            if hasattr(cell, 'value') and cell.value is not None:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column_letter].width = adjusted_width

if __name__ == "__main__":
    create_consolidated_test_suite()